"""Consolida os arquivos brutos de violência contra as mulheres no RS (data/raw/)
em tabelas "long" por município.
"""

from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.worksheet.worksheet import Worksheet

RAW_DIR = Path(__file__).resolve().parent.parent / "data" / "raw"
OUTPUT_DIR = Path(__file__).resolve().parent.parent / "outputs" / "tables"

# Único arquivo com estrutura diferente: um workbook consolidado cobrindo
# 2012-2017 (município x ano), em vez de um arquivo por ano (município x
# mês) como os demais.
CONSOLIDATED_FILE = "violencia_2012_2017.xlsx"

# Abas que não representam um crime com quebra por município e devem ser
# ignoradas ao montar as tabelas por município (a "Geral" é o total estadual
# somado entre categorias; a "Fórmulas..." é uma aba auxiliar oculta).
NON_CRIME_SHEETS = {"Geral", "Fórmulas - Demais indicadores"}

# Layout das abas de crime nos arquivos 2018+: município na coluna B,
# meses Jan-Dez nas colunas C-N, Total na coluna O. A linha de cabeçalho
# não está numa posição fixa (linha 3 na maioria dos anos, mas varia),
# por isso é localizada dinamicamente por _find_header_row.
MUNICIPIO_COL = 2  # B
MES_COLS = range(3, 15)  # C..N
TOTAL_COL = 15  # O

# Linha de total estadual ao final de cada aba de crime: marca o fim dos
# dados por município (o que vem depois é rodapé de fonte/observações).
STATE_TOTAL_LABEL = "Geral"

POPULACAO_FILE = "populacao_municipio_rs.csv"  # gerado por fetch_populacao.py

# Casos por 100 mil habitantes é a convenção mais comum em segurança
# pública no Brasil (ex.: taxa de homicídios). Note que isso usa população
# TOTAL, não só mulheres: o IBGE só quebra população por sexo no ano de
# Censo (não nas estimativas intercensitárias que cobrem os demais anos), e
# a DEE-RS tem população feminina por município mas só até 2021 -- não
# cobre 2022-2025. A planilha original da SSP calculava a taxa sobre
# população feminina ("mulheres vítimas por 10.000 habitantes", aba Geral
# do arquivo 2012-2017); a taxa aqui não é diretamente comparável a ela.
CASOS_POR_HABITANTES = 100_000


def normalize_crime_name(sheet_name: str) -> str:
    """Uniformiza a nomenclatura entre arquivos: 2012-2017 usa "Femicídio",
    2018+ usa "Feminicídio". O restante do nome é mantido como está.
    """
    return sheet_name.replace("Femicídio", "Feminicídio")


def open_workbook(path: Path) -> openpyxl.Workbook:
    """Abre a planilha em modo somente leitura, com data_only=True para obter
    o valor já calculado de células com fórmula (ex.: colunas "Total").
    """
    return openpyxl.load_workbook(path, read_only=True, data_only=True)


def get_crime_sheets(wb: openpyxl.Workbook) -> dict[str, Worksheet]:
    """Mapeia tipo_crime normalizado -> planilha (Worksheet), pulando as
    abas que não são de um crime específico (Geral, Fórmulas...).

    A ordem das abas "Feminicídio Tentado"/"Feminicídio Consumado" varia
    entre arquivos, então o mapeamento é sempre por nome, nunca por posição.
    """
    sheets: dict[str, Worksheet] = {}
    for name in wb.sheetnames:
        if name in NON_CRIME_SHEETS:
            continue
        sheets[normalize_crime_name(name)] = wb[name]
    return sheets


def _find_header_row(ws: Worksheet, max_search: int = 10) -> int:
    """Localiza a linha de cabeçalho procurando a célula "Município" na
    coluna B, em vez de assumir uma linha fixa (ela varia entre arquivos).

    Usa iter_rows (leitura sequencial) em vez de ws.cell(row=, column=):
    em modo read_only, cada chamada de ws.cell() reescaneia o XML da aba
    do início, o que é O(n) por célula e torna o parsing de uma aba de
    ~500 linhas absurdamente lento.
    """
    for row_idx, row in enumerate(ws.iter_rows(max_row=max_search), start=1):
        if row[MUNICIPIO_COL - 1].value == "Município":
            return row_idx
    raise ValueError(
        f"Cabeçalho 'Município' não encontrado nas primeiras {max_search} "
        f"linhas da aba {ws.title!r}"
    )


def _year_columns(ws: Worksheet, header_row: int) -> dict[int, int]:
    """Mapeia coluna -> ano a partir da linha de cabeçalho de uma aba do
    arquivo 2012-2017 (colunas de ano começam logo após a de município).

    O intervalo de anos varia por aba: "Feminicídio Tentado" só começa em
    2013 nesse arquivo (não era discriminado separadamente em 2012), as
    demais abas cobrem 2012-2017 inteiro. Por isso os anos são lidos do
    cabeçalho em vez de assumidos por posição fixa de coluna.
    """
    header = next(ws.iter_rows(min_row=header_row, max_row=header_row))
    year_cols = {}
    for col in range(MUNICIPIO_COL + 1, len(header) + 1):
        value = header[col - 1].value
        if not isinstance(value, int):
            break
        year_cols[col] = value
    return year_cols


def parse_annual_sheet_2012_2017(ws: Worksheet, tipo_crime: str) -> pd.DataFrame:
    """Extrai os casos anuais por município de uma aba de crime do arquivo
    consolidado 2012-2017 (município nas linhas, anos nas colunas).

    Ignora as colunas "Total", "% vítimas/total", "População de mulheres"
    e a taxa por 10.000 habitantes: não fazem parte do escopo da tabela
    anual por município e a taxa é um recorte de 2016, não uma série anual.
    A leitura para na linha "Geral" (total estadual), que não é um
    município.
    """
    header_row = _find_header_row(ws)
    year_cols = _year_columns(ws, header_row)
    records = []
    for row in ws.iter_rows(min_row=header_row + 1):
        municipio = row[MUNICIPIO_COL - 1].value
        if municipio is None:
            continue
        if municipio == STATE_TOTAL_LABEL:
            break
        for col, ano in year_cols.items():
            casos_total = row[col - 1].value
            records.append((municipio, ano, tipo_crime, casos_total))
    return pd.DataFrame(
        records, columns=["municipio", "ano", "tipo_crime", "casos_total"]
    )


def _year_from_filename(path: Path) -> int:
    """Extrai o ano de "violencia_<ano>.xlsx". Não é usado para o arquivo
    consolidado 2012-2017, que é tratado separadamente por CONSOLIDATED_FILE.
    """
    return int(path.stem.removeprefix("violencia_"))


def parse_monthly_sheet(ws: Worksheet, ano: int, tipo_crime: str) -> pd.DataFrame:
    """Extrai os casos mensais por município de uma aba de crime no formato
    dos arquivos 2018+ (município nas linhas, colunas Jan-Dez + Total).

    Meses ainda não ocorridos no ano corrente vêm como célula vazia (não
    zero) e são descartados, não convertidos em 0. A leitura para na linha
    "Geral" (total estadual), que não é um município.

    Em 2018 (Ameaça e Lesão Corporal) há uma linha extra "NÃO INFORMADO"
    antes da linha "Geral", para casos sem município identificado. Ela é
    mantida como um valor de município como outro qualquer, para não
    perder casos e preservar a reconciliação com o total estadual.

    Uma célula isolada (2018, Ameaça, linha "NÃO INFORMADO", outubro) usa
    "-" em vez de 0 para indicar ausência de casos; é normalizada para 0
    para manter a coluna "casos" numérica.
    """
    header_row = _find_header_row(ws)
    records = []
    for row in ws.iter_rows(min_row=header_row + 1):
        municipio = row[MUNICIPIO_COL - 1].value
        if municipio is None:
            continue
        if municipio == STATE_TOTAL_LABEL:
            break
        for mes, col in enumerate(MES_COLS, start=1):
            casos = row[col - 1].value
            if casos == "":
                continue
            if casos == "-":
                casos = 0
            records.append((municipio, ano, mes, tipo_crime, casos))
    return pd.DataFrame(
        records, columns=["municipio", "ano", "mes", "tipo_crime", "casos"]
    )


def parse_annual_sheet_year_file(ws: Worksheet, ano: int, tipo_crime: str) -> pd.DataFrame:
    """Extrai os casos anuais por município de uma aba de crime dos arquivos
    2018+, usando diretamente a coluna "Total" já calculada na planilha em
    vez de resomar os 12 meses (validado como idêntico à soma mensal em
    todos os arquivos, e mais direto). Mesmas regras de parada na linha
    "Geral" e de manutenção da linha "NÃO INFORMADO" de parse_monthly_sheet.
    """
    header_row = _find_header_row(ws)
    records = []
    for row in ws.iter_rows(min_row=header_row + 1):
        municipio = row[MUNICIPIO_COL - 1].value
        if municipio is None:
            continue
        if municipio == STATE_TOTAL_LABEL:
            break
        casos_total = row[TOTAL_COL - 1].value
        records.append((municipio, ano, tipo_crime, casos_total))
    return pd.DataFrame(
        records, columns=["municipio", "ano", "tipo_crime", "casos_total"]
    )


def load_monthly_long(raw_dir: Path = RAW_DIR) -> pd.DataFrame:
    """Monta a tabela mensal por município (2018-2026): município, ano, mês,
    tipo_crime, casos. O arquivo consolidado 2012-2017 não entra aqui — não
    tem quebra mensal por município (só a aba "Geral", em nível estadual).
    """
    frames = []
    year_files = sorted(
        p for p in raw_dir.glob("violencia_*.xlsx") if p.name != CONSOLIDATED_FILE
    )
    for path in year_files:
        ano = _year_from_filename(path)
        wb = open_workbook(path)
        for tipo_crime, ws in get_crime_sheets(wb).items():
            frames.append(parse_monthly_sheet(ws, ano, tipo_crime))
        wb.close()
    return pd.concat(frames, ignore_index=True)


def load_annual_long(raw_dir: Path = RAW_DIR) -> pd.DataFrame:
    """Monta a tabela anual por município (2012-2026): município, ano,
    tipo_crime, casos_total. Usa parse_annual_sheet_2012_2017 para o arquivo
    consolidado e parse_annual_sheet_year_file (coluna Total) para os demais.
    """
    frames = []

    wb = open_workbook(raw_dir / CONSOLIDATED_FILE)
    for tipo_crime, ws in get_crime_sheets(wb).items():
        frames.append(parse_annual_sheet_2012_2017(ws, tipo_crime))
    wb.close()

    year_files = sorted(
        p for p in raw_dir.glob("violencia_*.xlsx") if p.name != CONSOLIDATED_FILE
    )
    for path in year_files:
        ano = _year_from_filename(path)
        wb = open_workbook(path)
        for tipo_crime, ws in get_crime_sheets(wb).items():
            frames.append(parse_annual_sheet_year_file(ws, ano, tipo_crime))
        wb.close()

    return pd.concat(frames, ignore_index=True)


def load_annual_long_with_taxa(
    raw_dir: Path = RAW_DIR, tables_dir: Path = OUTPUT_DIR
) -> pd.DataFrame:
    """Junta a tabela anual por município com a população
    (outputs/tables/populacao_municipio_rs.csv, gerado por
    fetch_populacao.py) e calcula casos por 100 mil habitantes.

    A população só cobre 2012-2025 (2026 ainda não tem estimativa do IBGE)
    e não existe para a linha "NÃO INFORMADO" (não é um município real) —
    nessas linhas a taxa fica nula (NaN), não zero, para não confundir
    "sem população para calcular" com "taxa zero".
    """
    anual = load_annual_long(raw_dir)
    populacao = pd.read_csv(tables_dir / POPULACAO_FILE)

    df = anual.merge(populacao, on=["municipio", "ano"], how="left")
    df["taxa_por_100mil_hab"] = df["casos_total"] / df["populacao"] * CASOS_POR_HABITANTES
    return df


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    load_monthly_long().to_csv(
        OUTPUT_DIR / "violencia_mensal_municipio.csv", index=False
    )
    load_annual_long().to_csv(
        OUTPUT_DIR / "violencia_anual_municipio.csv", index=False
    )

    if (OUTPUT_DIR / POPULACAO_FILE).exists():
        load_annual_long_with_taxa().to_csv(
            OUTPUT_DIR / "violencia_anual_municipio_taxa.csv", index=False
        )
    else:
        print(
            f"Aviso: {POPULACAO_FILE} não encontrado em {OUTPUT_DIR} — rode "
            "src/fetch_populacao.py antes para gerar a tabela de taxas."
        )


if __name__ == "__main__":
    main()
