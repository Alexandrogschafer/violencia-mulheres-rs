"""Busca o vínculo município -> COREDE (Conselho Regional de Desenvolvimento)
e Região Funcional de Planejamento para os municípios do RS.

Diferente da malha do IBGE (fetch_malha_municipios.py) e das regiões
geográficas do IBGE (fetch_regioes_ibge.py), COREDEs são uma divisão
administrativa **estadual**, criada pela Lei Estadual 10.283/1994 -- não
existe no IBGE (nem na API de localidades, nem em nenhuma agregação do
SIDRA). A fonte estruturada oficial é o Atlas Socioeconômico do Rio Grande
do Sul (SEPLAG/RS), que disponibiliza a lista completa dos 497 municípios x
28 COREDEs x 9 Regiões Funcionais em uma planilha .xlsx, não uma API:

  https://atlassocioeconomico.rs.gov.br/upload/arquivos/202010/09172616-tabela-dos-municipios-por-corede-e-regiao-funcional-de-planejamento.xlsx

(encontrada a partir de https://atlassocioeconomico.rs.gov.br/conselhos-regionais-de-desenvolvimento-coredes)

Investigado e descartado antes de chegar aqui: dados.rs.gov.br (portal de
dados abertos do estado, busca por "corede"/"regionalizacao"/"desenvolvimento
regional" não retornou nenhum dataset com essa coluna) e fee.rs.gov.br
(fora do ar no momento da busca). Não confundir com a lista hardcoded de 13
municípios do COREDE Fronteira Oeste em notebooks/estudo_uruguaiana.ipynb --
aquela é só um recorte manual para o estudo de caso; esta planilha aqui é a
fonte completa e oficial, e bate com aquele recorte (mesmos 13 municípios)
como checagem de consistência.

Uso: python -m src.fetch_coredes
"""

import time
import urllib.request
from pathlib import Path

import openpyxl

from src.fetch_populacao import normaliza_municipio

URL_XLSX = (
    "https://atlassocioeconomico.rs.gov.br/upload/arquivos/202010/"
    "09172616-tabela-dos-municipios-por-corede-e-regiao-funcional-de-planejamento.xlsx"
)

RAW_PATH = Path(__file__).resolve().parent.parent / "data" / "raw" / "municipios_corede_rf.xlsx"
OUTPUT_PATH = (
    Path(__file__).resolve().parent.parent / "outputs" / "tables" / "municipio_corede.csv"
)

N_MUNICIPIOS_RS = 497
N_COREDES_RS = 28


def baixa_planilha(tentativas: int = 3) -> bytes:
    req = urllib.request.Request(URL_XLSX, headers={"User-Agent": "Mozilla/5.0"})
    ultimo_erro = None
    for tentativa in range(1, tentativas + 1):
        try:
            with urllib.request.urlopen(req, timeout=30) as resp:
                return resp.read()
        except Exception as erro:  # noqa: BLE001 - relançado após esgotar tentativas
            ultimo_erro = erro
            if tentativa < tentativas:
                time.sleep(2 * tentativa)
    raise RuntimeError(f"Falha ao baixar {URL_XLSX!r} após {tentativas} tentativas") from ultimo_erro


def parse_planilha(caminho: Path) -> list[dict]:
    """Extrai (codigo_ibge, municipio, corede, regiao_funcional) da planilha
    do Atlas Socioeconômico. Única aba, cabeçalho na linha 1, colunas
    CODIGO IBGE / MUNICÍPIO / COREDE / REGIÃO FUNCIONAL (confirmado
    inspecionando o arquivo antes de escrever este parser).
    """
    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    registros = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        codigo_ibge, municipio, corede, regiao_funcional = row[:4]
        if codigo_ibge is None:
            continue
        registros.append(
            {
                "codigo_ibge": str(int(codigo_ibge)),
                "municipio": normaliza_municipio(municipio),
                "corede": corede,
                "regiao_funcional": regiao_funcional,
            }
        )
    wb.close()
    return registros


def main() -> None:
    RAW_PATH.parent.mkdir(parents=True, exist_ok=True)
    if not RAW_PATH.exists():
        RAW_PATH.write_bytes(baixa_planilha())

    registros = parse_planilha(RAW_PATH)

    if len(registros) != N_MUNICIPIOS_RS:
        raise ValueError(
            f"Esperava {N_MUNICIPIOS_RS} municípios do RS, a planilha tem {len(registros)}."
        )
    codigos = {r["codigo_ibge"] for r in registros}
    if len(codigos) != N_MUNICIPIOS_RS:
        raise ValueError(
            f"código_ibge duplicado: {N_MUNICIPIOS_RS} municípios mas só "
            f"{len(codigos)} códigos únicos."
        )
    n_coredes = len({r["corede"] for r in registros})
    if n_coredes != N_COREDES_RS:
        raise ValueError(f"Esperava {N_COREDES_RS} COREDEs, a planilha tem {n_coredes}.")

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        f.write("codigo_ibge,municipio,corede,regiao_funcional\n")
        for r in registros:
            f.write(
                f"{r['codigo_ibge']},{r['municipio']},\"{r['corede']}\",{r['regiao_funcional']}\n"
            )

    n_rf = len({r["regiao_funcional"] for r in registros})
    print(
        f"Salvo em {OUTPUT_PATH} — {len(registros)}/{N_MUNICIPIOS_RS} municípios "
        f"validados, {n_coredes} COREDEs, {n_rf} regiões funcionais."
    )


if __name__ == "__main__":
    main()
