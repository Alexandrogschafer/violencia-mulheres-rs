"""Registra a proveniência dos arquivos brutos da SSP/RS (data/raw/violencia_*.xlsx).

Contexto: a SSP/RS revisa retroativamente os arquivos publicados (ex.: em
2026, a página mostra 2023 e 2024 "atualizados em janeiro de 2026", e 2021
"em dezembro de 2023") -- mas nada no repositório registrava qual versão
exata deste projeto usou, embora a informação já exista dentro dos próprios
arquivos: (1) a célula "Fonte: SIP/PROCERGS - Atualizado em <data>" na aba
Geral, escrita pela própria SSP/RS; e (2) o metadado OOXML
`docProps/core.xml:dcterms:modified`, atualizado pelo Excel a cada vez que a
SSP/RS salva o arquivo. As duas fontes concordam entre si em todos os
arquivos verificados manualmente antes de escrever este script -- ambas são
geradas pela SSP/RS e sobrevivem ao download e à renomeação para
`violencia_<ano>.xlsx` (nenhuma delas depende do nome do arquivo).

O que este script NÃO faz, de propósito: não infere nem uma data de
download real (só temos o mtime local do arquivo, que reflete quando este
projeto o *gravou em disco* -- pode coincidir com o download, mas também
pode ser de uma cópia/rsync/checkout posterior; por isso a coluna é rotulada
"aproximação", não "data de download"), não baixa nada da SSP/RS para
comparar com a versão atual do site, e não preenche a URL de origem (não
temos esse registro -- fica em branco, para preenchimento manual).

Uso: python -m src.verificar_proveniencia
"""

import hashlib
import re
from datetime import datetime
from pathlib import Path

import openpyxl

RAW_DIR = Path(__file__).resolve().parent.parent / "data" / "raw"
OUTPUT_PATH = (
    Path(__file__).resolve().parent.parent
    / "outputs"
    / "tables"
    / "proveniencia_dados_brutos.csv"
)

# Mesma extensão de "Atualizado em D/M/AAAA" usada pela SSP/RS em todos os
# arquivos verificados (dia/mês sem zero à esquerda, ano com 4 dígitos).
PADRAO_DATA_ATUALIZACAO = re.compile(r"Atualizado em (\d{1,2})/(\d{1,2})/(\d{4})")


def sha256_arquivo(caminho: Path) -> str:
    h = hashlib.sha256()
    with open(caminho, "rb") as f:
        for bloco in iter(lambda: f.read(1 << 20), b""):
            h.update(bloco)
    return h.hexdigest()


def extrai_celula_fonte(caminho: Path) -> str | None:
    """Busca, na aba 'Geral', a célula cujo texto começa com "Fonte:" --
    verbatim, sem normalizar. Retorna None se a aba não existir ou a célula
    não for encontrada (não inventa nem tenta adivinhar em outra aba)."""
    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    try:
        if "Geral" not in wb.sheetnames:
            return None
        ws = wb["Geral"]
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.strip().startswith("Fonte:"):
                    return cell.value.strip()
        return None
    finally:
        wb.close()


def extrai_modified_ooxml(caminho: Path) -> datetime | None:
    """dcterms:modified de docProps/core.xml, já como datetime (openpyxl
    parseia isso em wb.properties.modified -- não precisa ler o XML manualmente)."""
    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    try:
        return wb.properties.modified
    finally:
        wb.close()


def normaliza_data_atualizacao(fonte_verbatim: str | None) -> str | None:
    """Extrai D/M/AAAA de dentro do texto verbatim da célula "Fonte:" e
    normaliza para ISO (AAAA-MM-DD). O texto verbatim continua registrado à
    parte (coluna fonte_celula_verbatim) -- esta função é só conveniência de
    leitura, nunca a evidência em si."""
    if not fonte_verbatim:
        return None
    m = PADRAO_DATA_ATUALIZACAO.search(fonte_verbatim)
    if not m:
        return None
    dia, mes, ano = (int(x) for x in m.groups())
    return f"{ano:04d}-{mes:02d}-{dia:02d}"


def processa_arquivo(caminho: Path) -> dict:
    fonte_verbatim = extrai_celula_fonte(caminho)
    modified_ooxml = extrai_modified_ooxml(caminho)
    mtime_local = datetime.fromtimestamp(caminho.stat().st_mtime)

    aviso = None
    if fonte_verbatim is None:
        aviso = "AVISO: célula 'Fonte:' não encontrada na aba Geral -- data de atualização declarada indisponível."

    return {
        "arquivo": caminho.name,
        "sha256": sha256_arquivo(caminho),
        "fonte_celula_verbatim": fonte_verbatim or "",
        "fonte_data_atualizacao_iso": normaliza_data_atualizacao(fonte_verbatim) or "",
        "ooxml_dcterms_modified_iso": modified_ooxml.isoformat() if modified_ooxml else "",
        "mtime_local_aproximacao_download": mtime_local.isoformat(),
        "url_origem": "",  # preenchimento manual -- ver docstring do módulo
        "aviso": aviso or "",
    }


def main() -> None:
    arquivos = sorted(RAW_DIR.glob("violencia_*.xlsx"))
    if not arquivos:
        raise FileNotFoundError(f"Nenhum violencia_*.xlsx encontrado em {RAW_DIR}")

    registros = [processa_arquivo(caminho) for caminho in arquivos]

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    colunas = [
        "arquivo",
        "sha256",
        "fonte_celula_verbatim",
        "fonte_data_atualizacao_iso",
        "ooxml_dcterms_modified_iso",
        "mtime_local_aproximacao_download",
        "url_origem",
        "aviso",
    ]
    with open(OUTPUT_PATH, "w", encoding="utf-8", newline="") as f:
        f.write(",".join(colunas) + "\n")
        for r in registros:
            linha = [str(r[c]).replace('"', '""') for c in colunas]
            linha = [f'"{v}"' if ("," in v or '"' in v) else v for v in linha]
            f.write(",".join(linha) + "\n")

    n_com_aviso = sum(1 for r in registros if r["aviso"])
    print(f"Salvo em {OUTPUT_PATH} — {len(registros)} arquivo(s) processados.")
    if n_com_aviso:
        print(f"AVISO: {n_com_aviso} arquivo(s) sem célula 'Fonte:' encontrada (ver coluna 'aviso' no CSV):")
        for r in registros:
            if r["aviso"]:
                print(f"  - {r['arquivo']}")
    else:
        print("Célula 'Fonte:' encontrada em todos os arquivos.")


if __name__ == "__main__":
    main()
